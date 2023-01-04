from openpyxl import workbook, load_workbook

FILE_DIR = 'A:\Master File\Key - AIFMD Manual Data.xlsm'


wb = load_workbook(FILE_DIR)
ws = wb.active
EURUSD_RATE = ws['B5'].value

# Q4 2021:
# EURUSD_RATE = 1.1326  # 31/12/2021 on the ecb website
# https://www.ecb.europa.eu/stats/policy_and_exchange_rates/euro_reference_exchange_rates/html/eurofxref-graph-usd.en.html

# 118
MainBeneficialOwnersRate_alto = round(ws['B8'].value*100, 2)  # Beneficially owned percentage by top 5 beneficial owners
MainBeneficialOwnersRate_neutral = round(ws['C8'].value*100, 2)  # Beneficially owned percentage by top 5 beneficial owners

# 137
AnnualInvestmentReturnRate_alto = round(ws['B9'].value*100, 2)   # Expected annual investment return in %
AnnualInvestmentReturnRate_neutral = round(ws['C9'].value*100, 2)   # Expected annual investment return in %

# 219-230 Gross investment return in % for the 3 month of the quarter (2 decimals)
GrossInvestmentReturnsRate_alto = [round(ws['B12'].value, 2), round(ws['C12'].value, 2), round(ws['D12'].value, 2)]
GrossInvestmentReturnsRate_neutral = [round(ws['B19'].value, 2), round(ws['C19'].value, 2), round(ws['D19'].value, 2)]

# 231-242 Percentage of net investment returns in % for the 3 month of the quarter (2 decimals)
NetInvestmentReturnsRate_alto = [round(ws['B13'].value, 2), round(ws['C13'].value, 2), round(ws['D13'].value, 2)]
NetInvestmentReturnsRate_neutral = [round(ws['B20'].value, 2), round(ws['C20'].value, 2), round(ws['D20'].value, 2)]

# 243-254 NAV change in % for the 3 month of the quarter (2 decimals)
NAVChangeRate_alto = [round(ws['B14'].value, 2), round(ws['C14'].value, 2), round(ws['D14'].value, 2)]
NAVChangeRate_neutral = [round(ws['B21'].value, 2), round(ws['C21'].value, 2), round(ws['D21'].value, 2)]

# 255-266 Subscription amount in $ for the 3 month of the quarter (0 decimals)
Subscription_alto = [round(ws['B15'].value, 0), round(ws['C15'].value, 0), round(ws['D15'].value, 0)]
Subscription_neutral = [round(ws['B22'].value, 0), round(ws['C22'].value, 0), round(ws['D22'].value, 0)]

# 267-278 Redemption amount in $ for the 3 month of the quarter (0 decimals)
Redemption_alto = [int(ws['B16'].value), int(ws['C16'].value), int(ws['D16'].value)]
Redemption_neutral = [int(ws['B23'].value), int(ws['C23'].value), int(ws['D23'].value)]

# 279 Results of stress tests performed in accordance with point(b) of Article 15(3)
StressTestsResultArticle15_alto = ws['B26'].value
StressTestsResultArticle15_neutral = ws['H26'].value

# 280 Results of stress tests performed in accordance with the second subparagraph of Article 16(1)
StressTestsResultArticle16_alto = ws['B27'].value
StressTestsResultArticle16_neutral = ws['H27'].value









